import json
import requests
import yaml

from openpyxl import load_workbook

from login_cloud import import_login, read_config, get_env_url

# config = read_config()
# user1 = config["test_env_username1"]

# import_login = import_login(user1)
# headers = import_login.headers()
# crop_id = import_login.crop_id()
# org_id = import_login.org_id()
env_url =get_env_url()

def read_config():
    conf_path = r"D:\git\ui_test\import_branch\branch_account.yaml"
    with open(conf_path, encoding='utf-8',) as f:
        env_config = yaml.load(f.read(), Loader=yaml.SafeLoader)
    return env_config
# config = read_config()
# username = config["test_env_username"]
#
# print("username:", username)
# sheet1 = config["sheet1"]
# sheet2 = config["sheet2"]
# sheet3 = config["sheet3"]
# sheet4 = config["sheet4"]

class AddShop:
    
    def __init__(self, crop_id,org_id,headers,username,sheet_index):
        self.crop_id = crop_id
        self.org_id = org_id
        self.headers =headers
        self.username = username
        self.sheet_index = sheet_index

    def read_branch(self):
        # 加载工作簿
        workbook = load_workbook(filename=r"D:\git\ui_test\import_branch\import_data.xlsx")
        print("all sheets:",workbook.sheetnames)

        # # 4个账号：悸动   good   luckin  snow
        # if self.username == "悸动烧仙草":
        #     sheet_index = 0
        # elif self.username == "good":
        #     sheet_index = 1
        # elif self.username == "luckin":
        #     sheet_index = 2
        # elif self.username == "snow":
        #     sheet_index = 3
        # else:
        #     print("配置文件中的账号不对")
        #
        # print("read branch 的user",self.username)


        # 选择工作表，这里假设是第一个工作表
        print("当前打开的sheet index是*****************************************************：",self.sheet_index)
        sheet = workbook.worksheets[self.sheet_index]
        # sheet = workbook.worksheets[0]  # 或者使用 sheet = workbook['Sheet1'] 来指定工作表名

        # sheet = workbook[sheet_name]
        # 初始化一个空字典来存储每列的数据
        column_data = {}

        # 遍历每一行
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，所以从第二行开始读取
            for col_idx, value in enumerate(row, start=1):  # enumerate从1开始计数，以匹配Excel的列号
                if col_idx not in column_data:
                    column_data[col_idx] = []
                column_data[col_idx].append(value)

            # 现在column_data字典包含了每一列的数据
        # 如果你想将每列的数据转换成列表，可以直接从字典中获取
        # 例如，获取第一列的数据
        first_column_data = column_data[1]
        # 打印所有列的数据
        return_all = []
        for col_idx, data in column_data.items():
            data = list(filter(lambda x: x is not None, data))
            return_all.append(data)
        # print("return_all:", return_all)
        # print("return_all的类型", type(return_all))
        # print("return_all的长度", len(return_all))

        first_row_values = []
        # 每列第一行的值
        for col in sheet.iter_cols(min_row=1, max_row=1, values_only=False):
            # 由于我们只需要第一行的值，所以只取第一个单元格（实际上是唯一的单元格，因为我们限制了max_row=1）
            cell = col[0]
            # 将单元格的值添加到列表中
            first_row_values.append(cell.value)

            # 输出包含每一列第一行值的列表
        print("first_row_values", first_row_values)

        # 总列数
        max_cols = sheet.max_column
        print("max_cols:", max_cols)

        # 总行数
        max_rows = sheet.max_row
        print("max_rows:", max_rows)

        max_rows_per_column = []
        # 遍历工作表中的所有列（这里我们假设我们想要检查所有列）
        # 注意：在openpyxl中，列是通过其索引（从1开始）来访问的，但我们可以使用ws.iter_cols()来迭代列
        for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            # 初始化该列的最大行数为0
            max_row = 0
            # 遍历该列中的每一行
            for cell in col:
                # 如果单元格不为空，则更新最大行数
                if cell.value is not None:
                    max_row = cell.row
                    # 将该列的最大行数添加到列表中
            max_rows_per_column.append(max_row)

            # 输出包含每列最大行数的列表
        print("max_rows_per_column", max_rows_per_column)

        # 返回每列的行数、每列第一行的值、每列的值
        return first_row_values, return_all

    # 获取所有“分支”类型的机构名称的机构id
    def get_type_branchs(self):
        get_org_id_url = f'{env_url}/cuteview/org/get'
        body = {
            "cropId": self.crop_id,
        }
        resp = requests.get(url=get_org_id_url, params=body, headers=self.headers)

        # 获取所有的机构类型和机构名称
        resp = resp.json()
        # print("resp", resp)

        org_types = [item["type"] for item in resp["data"]["allChildrenOrg"]]
        # org_names = [item["name"] for item in resp["data"]["allChildrenOrg"]]
        # name_type = {k: v for k, v in zip(org_names,org_types )}
        self.org_ids = [item["id"] for item in resp["data"]["allChildrenOrg"]]
        ids_type = {k: v for k, v in zip(self.org_ids, org_types)}

        # print("ids_type", ids_type)

        branch_type_ids = []
        shop_type_ids = []
        for name, type in ids_type.items():
            if type == 2:
                branch_type_ids.append(name)
            else:
                shop_type_ids.append(name)

            # print(name)
            # print(type)
            # # print("names",names)

        print("接口函数中的self.org_ids：-***********************---",self.org_ids)
        print("接口函数中的branch_type_ids:",branch_type_ids)
        print("接口函数中的shop_type_ids:", shop_type_ids)
        return self.org_ids, branch_type_ids, shop_type_ids

    def get_root_node_id(self):
        get_org_id_url = f'{env_url}/cuteview/org/get'
        body = {
            "cropId": self.crop_id,
        }
        resp = requests.get(url=get_org_id_url, params=body, headers=self.headers)

        # 获取所有的机构类型和机构名称
        resp = resp.json()
        # print("resp", resp)
        root_node_id = resp["data"]["rootNode"]["id"]
        print("root_node_id：----", root_node_id)
        return root_node_id

    def clear_all_org(self):
        ids, _, _ = self.get_type_branchs()
        # print("所有的child ids:", ids)

        iter_times = len(ids)

        for id in ids:
            # print("child id:", id)
            clear_base_url = f'{env_url}/cuteview/org/delete/{id}'
            # print("clear_base_url-------:", clear_base_url)

            # for i in range(0, iter_times):

            # print("clear_base_url:", clear_base_url)
            body = {
                "cropId": self.crop_id,
            }
            resp = requests.post(url=clear_base_url, headers=self.headers, data=json.dumps(body))
            # print(resp.json())

            # logger.info("success clear org")
            # print("执行清理多少次")

    def add_each_branch(self):
        # 添加所有分支
        first_row_values, _ = self.read_branch()
        parent_id = self.get_root_node_id()
        add_url = f'{env_url}/cuteview/org/create'
        for branch in first_row_values:
            # print("all branch:", branch)
            # print("branch-type:", type(branch))

            data = {
                "contact": "",
                "cropId": self.crop_id,
                "id": "",
                "level": 2,
                "name": branch,
                "outerCode": "",
                "parentId": parent_id,
                "phone": "",
                "type": 2,
                "businessStartTime": "",
                "businessEndTime": "",
                "orgTypeId": "",
                "orgLevelId": ""
            }

            resp = requests.post(url=add_url, headers=self.headers, data=json.dumps(data))

            # print("resp:", resp.json())

    def add_each_shop(self):
        # 添加所有门店
        _, ids, _ = self.get_type_branchs()

        # print("ids++++++++++++++：", ids)
        # for id in ids:
        #     print("每次的id的值是:",id)
        _, return_all = self.read_branch()

        ii = len(ids)
        for id in ids:
            # print("id的值是:",id)

            indices = [i for i, value in enumerate(ids) if value == id]
            # print("indices:", indices)
            # print(indices[0])
            # print("return_all类型:", type(return_all))
            # print("return_all.index(indices)", return_all[indices[0]])
            for shop in return_all[indices[0]]:
                # print("each:", shop)
                # for i in return_all.index(indices):
                #
                #     # print("list中的值:", i)
                #     for j in i:
                #         # print("j", j)
                #         pass

                add_url = f'{env_url}/cuteview/org/create'
                data = {
                    "contact": "",
                    "cropId": self.crop_id,
                    "id": "",
                    "level": 2,
                    "name": shop,
                    "outerCode": "",
                    "parentId": id,
                    "phone": "",
                    "type": 1,
                    "businessStartTime": "",
                    "businessEndTime": "",
                    "orgTypeId": "",
                    "orgLevelId": ""
                }
                resp = requests.post(url=add_url, headers=self.headers, data=json.dumps(data))
                # print("resp:", resp.json())

#
# if __name__ =="__main__":
#     add_shop =AddShop(crop_id,org_id,headers,username)
#     add_shop.read_branch()


