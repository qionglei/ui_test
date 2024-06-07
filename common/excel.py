import openpyxl


def read_excel(excel_dir, sheet_name):
    """读取excel"""

    # 加载目录
    ex = openpyxl.load_workbook(excel_dir)

    # 获取sheet页
    sheet = ex[sheet_name]

    # 打印表最大行和列
    # print(sheet.max_row,sheet.max_column)
    # print(sheet.cell(2,1).value)
    # 循环行和列
    sheet_list = []
    for row in range(2, sheet.max_row + 1):
        row_list = []
        for col in range(1, sheet.max_column + 1):
            row_list.append(sheet.cell(row, col).value)
        sheet_list.append(row_list)
    return sheet_list

# if __name__ == '__main__':
#     print("excel.py处运行的main函数")
#     read_excel('../data/data.xlsx','login')
